import logging
from openpyxl import load_workbook
from app.schemas import GeometryItem

logger = logging.getLogger(__name__)

def load_geometry_items(xlsx_path: str) -> list[GeometryItem]:
    logger.info("load_geometry_items called xlsx_path=%s", xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb.active
    items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[1]:
            continue
        items.append(GeometryItem(name=str(row[0]), WKT=str(row[1])))

    logger.info("load_geometry_items response_count=%s", len(items))
    return items