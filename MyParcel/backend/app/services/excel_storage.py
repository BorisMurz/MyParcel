import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)

def save_geometry(file_name: str, wkt: str, xlsx_path: str) -> None:
    logger.info("save_geometry called file_name=%s xlsx_path=%s", file_name, xlsx_path)

    path = Path(xlsx_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "geometry"
        ws.append(["Наименование файла", "Описание геометрии"])

    ws.append([file_name, wkt])
    wb.save(path)

    logger.info("save_geometry saved successfully")